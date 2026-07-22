#!/usr/bin/env python3
"""
xlsxspec.py — Convert Excel (.xlsx) design-spec sheets to rg-friendly text.

Output formats (--format):
  md     GFM tables. Human/LLM readable. NOT ideal for rg.
  tsv    `sheet<TAB>cell<TAB>value` per line. Best for rg/grep/index.
  jsonl  one JSON object per line. rg-able AND jq-queryable.
  kv     like tsv but wrapped with `# sheet` markdown headers (legacy).

Why these beat a plain markdown table for ripgrep:
  - one value per line, each line self-describing (sheet + cell + value)
  - no `|`/`---`/padding noise, no escaped pipes
  - tab/JSON delimiters never collide with Japanese text

Grid-mode (md) improvements over a naive converter:
  - real bounding-box detection (ignores ghost 1000x26 / 220x68 dims)
  - GLOBAL empty-column drop (blanks *between* data, not just trailing)
  - optional --fill-merge to propagate a merged value across its range
"""

import argparse
import json
import re
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Optional

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
except ImportError:
    sys.exit("openpyxl not found: pip install openpyxl")


# Force UTF-8 on stdout/stderr so Japanese sheet names / paths never crash the
# tool on consoles with a legacy encoding (e.g. Windows cp1252). Self-sufficient
# regardless of whether the caller exported PYTHONUTF8/PYTHONIOENCODING.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="backslashreplace")
    except (AttributeError, ValueError):
        pass


# ─── Escaping (only for md) ─────────────────────────────────────────────────────
def _escape_md(text: str) -> str:
    text = text.replace("\\", "\\\\").replace("|", "\\|")
    return text.replace("\n", " ").replace("\r", "")


def _clean(text: str) -> str:
    """For tsv/jsonl: strip line breaks and tabs, no markdown escaping."""
    return text.replace("\n", " ").replace("\r", "").replace("\t", " ").strip()


# ─── Value formatting ───────────────────────────────────────────────────────────
def _fmt_value(value, number_format: str) -> str:
    nf = (number_format or "").upper()
    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M") if (value.hour or value.minute) else value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d")
    if "%" in nf:
        try:
            return f"{float(value)*100:.2f}%"
        except (TypeError, ValueError):
            pass
    for sym in ("$", "€", "£", "¥", "₫"):
        if sym in (number_format or ""):
            try:
                return f"{sym}{float(value):,.2f}"
            except (TypeError, ValueError):
                pass
    if isinstance(value, float):
        return str(int(value)) if value == int(value) else f"{value:,}"
    if isinstance(value, int):
        return f"{value:,}"
    return str(value)


def _raw_cell_value(cell) -> Optional[str]:
    """Return cleaned string for a cell, or None if empty. No markdown styling."""
    value = cell.value
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, datetime, date)):
        return _fmt_value(value, cell.number_format)
    return str(value).strip()


# ─── Merge handling ─────────────────────────────────────────────────────────────
def _build_merge_value_map(ws, fill: bool) -> dict:
    m = {}
    for rng in ws.merged_cells.ranges:
        anchor = ws.cell(rng.min_row, rng.min_col).value if fill else None
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if r == rng.min_row and c == rng.min_col:
                    continue
                m[(r, c)] = anchor if fill else ""
    return m


def _merge_skip_set(ws) -> set:
    """Non-anchor cells of every merge range (so we emit each merge once)."""
    skip = set()
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if not (r == rng.min_row and c == rng.min_col):
                    skip.add((r, c))
    return skip


def _real_bbox(ws):
    maxr = maxc = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                if cell.row > maxr:
                    maxr = cell.row
                if cell.column > maxc:
                    maxc = cell.column
    return maxr, maxc


# ─── md (grid) helpers ──────────────────────────────────────────────────────────
def _cell_text_md(cell, merge_map, hyperlinks) -> str:
    if isinstance(cell, MergedCell):
        v = merge_map.get((cell.row, cell.column), "")
        return "" if v in (None, "") else _escape_md(str(v).strip())
    value = cell.value
    if value is None:
        return ""
    if hyperlinks and cell.hyperlink and cell.hyperlink.target:
        return f"[{_escape_md(str(value).strip())}]({cell.hyperlink.target.strip()})"
    text = _fmt_value(value, cell.number_format) if isinstance(value, (int, float, datetime, date)) else str(value).strip()
    text = _escape_md(text)
    if cell.font:
        b, i = cell.font.bold, cell.font.italic
        if b and i:
            text = f"***{text}***"
        elif b:
            text = f"**{text}**"
        elif i:
            text = f"*{text}*"
    return text


def _drop_empty_cols(rows):
    if not rows:
        return rows
    n = max(len(r) for r in rows)
    rows = [r + [""] * (n - len(r)) for r in rows]
    keep = [i for i in range(n) if any(r[i].strip() for r in rows)]
    return [[r[i] for i in keep] for r in rows]


def _drop_empty_rows(rows):
    return [r for r in rows if any(c.strip() for c in r)]


def _render_table(rows, max_w):
    if not rows:
        return ""
    n = len(rows[0])
    rows = [r + [""] * (n - len(r)) for r in rows]
    if max_w:
        rows = [[c[:max_w] + ("…" if len(c) > max_w else "") for c in r] for r in rows]
    widths = [max(max((len(r[i]) for r in rows), default=3), 3) for i in range(n)]

    def line(cells):
        return "| " + " | ".join(cells[i].ljust(widths[i]) for i in range(n)) + " |"

    out = [line(rows[0]), "| " + " | ".join("-" * widths[i] for i in range(n)) + " |"]
    out += [line(r) for r in rows[1:]]
    return "\n".join(out)


def convert_sheet_md(ws, *, header_row, max_col_width, hyperlinks, fill_merge, verbose) -> str:
    if verbose:
        print(f"  -> '{ws.title}' declared {ws.max_row}x{ws.max_column}")
    maxr, maxc = _real_bbox(ws)
    if maxr == 0:
        return f"*Sheet `{ws.title}` has no data.*\n"
    merge_map = _build_merge_value_map(ws, fill=fill_merge)
    rows = [[_cell_text_md(c, merge_map, hyperlinks) for c in row]
            for row in ws.iter_rows(min_row=header_row, max_row=maxr, max_col=maxc)]
    rows = _drop_empty_cols(_drop_empty_rows(rows))
    rows = _drop_empty_rows(rows)
    if not rows:
        return f"*Sheet `{ws.title}` has no data.*\n"
    return _render_table(rows, max_col_width) + "\n"


# ─── line-oriented record extraction (shared by tsv/jsonl/kv) ───────────────────
def iter_records(ws):
    """Yield (cell_coord, row, col, value) for every non-empty, non-duplicate cell."""
    maxr, maxc = _real_bbox(ws)
    if maxr == 0:
        return
    skip = _merge_skip_set(ws)
    for row in ws.iter_rows(min_row=1, max_row=maxr, max_col=maxc):
        for cell in row:
            if (cell.row, cell.column) in skip:
                continue
            v = _raw_cell_value(cell)
            if v is None:
                continue
            yield cell.coordinate, cell.row, cell.column, _clean(v)


def convert_sheet_tsv(ws, sheet_label) -> str:
    lines = [f"{sheet_label}\t{coord}\t{val}" for coord, _, _, val in iter_records(ws)]
    return "\n".join(lines) + ("\n" if lines else "")


def convert_sheet_jsonl(ws, sheet_label, source) -> str:
    out = []
    for coord, r, c, val in iter_records(ws):
        out.append(json.dumps(
            {"source": source, "sheet": sheet_label, "cell": coord,
             "row": r, "col": c, "value": val},
            ensure_ascii=False))
    return "\n".join(out) + ("\n" if out else "")


# ─── Front matter ──────────────────────────────────────────────────────────────
def _front_matter(src, sheets):
    mtime = datetime.fromtimestamp(src.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    return "\n".join([
        "---", f'source: "{src.name}"',
        f'converted: "{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}"',
        f'modified: "{mtime}"', f"sheets: [{', '.join(sheets)}]", "---", "",
    ])


def _safe(name):
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name)


# ─── Driver ──────────────────────────────────────────────────────────────────────
EXT = {"md": ".md", "kv": ".md", "tsv": ".tsv", "jsonl": ".jsonl"}


def convert(input_path, output_path=None, *, fmt="md", sheet_names=None, split_sheets=False,
            header_row=1, max_col_width=None, hyperlinks=True, front_matter=True,
            fill_merge=False, verbose=False):
    src = Path(input_path).expanduser().resolve()
    if not src.exists():
        sys.exit(f"Not found: {src}")
    wb = load_workbook(src, data_only=True)
    targets = sheet_names or wb.sheetnames
    missing = [s for s in targets if s not in wb.sheetnames]
    if missing:
        sys.exit(f"Sheets not found: {missing}")
    out_base = Path(output_path) if output_path else src.with_suffix("")

    # front matter only makes sense for the markdown-flavoured outputs
    use_fm = front_matter and fmt in ("md", "kv")

    def render_sheet(name):
        if fmt == "md":
            return f"# {name}\n\n" + convert_sheet_md(
                wb[name], header_row=header_row, max_col_width=max_col_width,
                hyperlinks=hyperlinks, fill_merge=fill_merge, verbose=verbose)
        if fmt == "kv":
            return f"# {name}\n\n" + convert_sheet_tsv(wb[name], name)
        if fmt == "tsv":
            return convert_sheet_tsv(wb[name], name)
        if fmt == "jsonl":
            return convert_sheet_jsonl(wb[name], name, src.name)
        raise ValueError(fmt)

    def build(names):
        parts = [_front_matter(src, names)] if use_fm else []
        for name in names:
            parts.append(render_sheet(name))
            if fmt in ("md", "kv"):
                parts.append("")
        return "\n".join(parts) if fmt in ("md", "kv") else "".join(parts)

    if split_sheets:
        out_base.mkdir(parents=True, exist_ok=True)
        for name in targets:
            f = out_base / f"{_safe(name)}{EXT[fmt]}"
            f.write_text(build([name]), encoding="utf-8")
            print(f"wrote {f}")
    else:
        out_file = Path(str(out_base) + EXT[fmt]) if not str(out_base).endswith(EXT[fmt]) else out_base
        out_file.write_text(build(targets), encoding="utf-8")
        print(out_file)


def main():
    p = argparse.ArgumentParser(prog="xlsxspec", description="XLSX design-spec -> rg-friendly text")
    p.add_argument("input")
    p.add_argument("-o", "--output")
    p.add_argument("-f", "--format", choices=["md", "tsv", "jsonl", "kv"], default="md",
                   help="output format (default md). tsv = best for rg.")
    p.add_argument("--sheets", nargs="+")
    p.add_argument("--split", action="store_true", help="one file per sheet")
    p.add_argument("--header-row", type=int, default=1, help="md mode only")
    p.add_argument("--max-col-width", type=int, help="md mode only")
    p.add_argument("--fill-merge", action="store_true",
                   help="md mode: propagate merged value across its whole range")
    p.add_argument("--no-hyperlinks", action="store_true")
    p.add_argument("--no-front-matter", action="store_true")
    p.add_argument("-v", "--verbose", action="store_true")
    a = p.parse_args()
    convert(a.input, a.output, fmt=a.format, sheet_names=a.sheets, split_sheets=a.split,
            header_row=a.header_row, max_col_width=a.max_col_width,
            hyperlinks=not a.no_hyperlinks, front_matter=not a.no_front_matter,
            fill_merge=a.fill_merge, verbose=a.verbose)


if __name__ == "__main__":
    main()
