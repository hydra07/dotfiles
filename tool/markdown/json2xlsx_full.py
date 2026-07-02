import json
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

from _xlsx_common import decode_val, setup_utf8_stdout, thin_border

setup_utf8_stdout()


def deserialize_workbook(json_path, xlsx_path):
    print(f"⏳ Đang phục hồi Excel từ file JSON: {json_path}...")
    with open(json_path, "r", encoding="utf-8") as f:
        wb_data = json.load(f)

    wb = openpyxl.Workbook()
    # Xoá sheet mặc định
    default_sheet = wb.active
    wb.remove(default_sheet)

    border_all = thin_border()

    for sheet_name, sheet_info in wb_data["sheets"].items():
        print(f"  -> Đang xây dựng lại sheet: {sheet_name}")
        ws = wb.create_sheet(title=sheet_name)
        
        # Grid lines
        if ws.views.sheetView:
            ws.views.sheetView[0].showGridLines = sheet_info.get("show_grid_lines", True)

        # Cài đặt độ rộng cột
        for col_idx_str, width in sheet_info.get("col_widths", {}).items():
            ws.column_dimensions[get_column_letter(int(col_idx_str))].width = width

        # Cài đặt độ cao dòng
        for row_idx_str, height in sheet_info.get("row_heights", {}).items():
            ws.row_dimensions[int(row_idx_str)].height = height

        # Ghi các ô dữ liệu và định dạng
        for cell_data in sheet_info["cells"]:
            r = cell_data["r"]
            c = cell_data["c"]
            val = cell_data.get("val")
            formula = cell_data.get("formula")
            bg = cell_data.get("bg")
            font_info = cell_data.get("font")
            align_info = cell_data.get("align")
            border_info = cell_data.get("border")
            num_fmt = cell_data.get("num_fmt")
            link = cell_data.get("link")
            comment_text = cell_data.get("comment")

            cell = ws.cell(row=r, column=c)
            cell.border = border_all

            # Ghi giá trị hoặc công thức
            if formula:
                cell.value = formula
            else:
                cell.value = decode_val(val)

            # Phục hồi màu nền
            if bg:
                # openpyxl cần ARGB hex 8 kí tự. Nếu chỉ có 6 kí tự thì thêm FF ở đầu
                hex_color = bg if len(bg) == 8 else "FF" + bg
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

            # Phục hồi Font
            if font_info:
                f_color = font_info["color"]
                f_hex = None
                if f_color:
                    f_hex = f_color if len(f_color) == 8 else "FF" + f_color
                
                cell.font = Font(
                    name=font_info.get("name"),
                    size=font_info.get("size"),
                    bold=font_info.get("bold"),
                    italic=font_info.get("italic"),
                    color=f_hex
                )

            # Phục hồi Căn lề
            if align_info:
                cell.alignment = Alignment(
                    horizontal=align_info.get("horizontal"),
                    vertical=align_info.get("vertical"),
                    wrap_text=align_info.get("wrap_text")
                )

            # Phục hồi Border
            if border_info:
                cell.border = Border(
                    left=Side(style=border_info.get("left")) if border_info.get("left") else None,
                    right=Side(style=border_info.get("right")) if border_info.get("right") else None,
                    top=Side(style=border_info.get("top")) if border_info.get("top") else None,
                    bottom=Side(style=border_info.get("bottom")) if border_info.get("bottom") else None
                )
            
            # Phục hồi Định dạng số
            if num_fmt:
                cell.number_format = num_fmt
                
            # Phục hồi Hyperlink
            if link:
                cell.hyperlink = link
                
            # Phục hồi Comment
            if comment_text:
                cell.comment = Comment(comment_text, "Author")

        # Khôi phục các vùng ô bị gộp
        for range_str in sheet_info.get("merged_ranges", []):
            try:
                ws.merge_cells(range_str)
            except Exception as e:
                # Tránh lỗi nếu trùng lặp
                pass

    wb.save(xlsx_path)
    print(f"✅ Đã phục hồi thành công toàn bộ workbook tại: {xlsx_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Sử dụng: python json2xlsx_full.py <input.json> <output.xlsx>")
        sys.exit(1)

    deserialize_workbook(sys.argv[1], sys.argv[2])
