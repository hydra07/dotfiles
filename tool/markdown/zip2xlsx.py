import json
import sys
import csv
import io
import zipfile
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

from _xlsx_common import setup_utf8_stdout

setup_utf8_stdout()


def deserialize_from_zip(zip_path, xlsx_path):
    print(f"⏳ Đang phục hồi Excel từ gói dữ liệu ZIP: {zip_path}...")

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    with zipfile.ZipFile(zip_path, 'r') as zf:
        # 1. Đọc metadata.json
        with zf.open("metadata.json") as f:
            metadata = json.load(f)

        # 2. Xử lý từng sheet
        for sheet_name, sheet_info in metadata["sheets"].items():
            print(f"  -> Xây dựng lại sheet: {sheet_name}")
            ws = wb.create_sheet(title=sheet_name)

            if ws.views.sheetView:
                ws.views.sheetView[0].showGridLines = sheet_info.get("show_grid_lines", True)

            # Phục hồi độ rộng cột / chiều cao dòng
            for col_idx_str, width in sheet_info.get("col_widths", {}).items():
                ws.column_dimensions[get_column_letter(int(col_idx_str))].width = width
            for row_idx_str, height in sheet_info.get("row_heights", {}).items():
                ws.row_dimensions[int(row_idx_str)].height = height

            # Phục hồi các vùng ô bị gộp
            for range_str in sheet_info.get("merged_ranges", []):
                try:
                    ws.merge_cells(range_str)
                except Exception:
                    pass

            # Đọc CSV cho dữ liệu Grid
            csv_filename = f"sheets/{sheet_name}.csv"
            if csv_filename in zf.namelist():
                with zf.open(csv_filename) as f:
                    # Chuyển bytes thành string để đọc CSV
                    csv_text = f.read().decode('utf-8-sig')
                    reader = csv.reader(io.StringIO(csv_text))
                    
                    for r_idx, row in enumerate(reader, 1):
                        for c_idx, val in enumerate(row, 1):
                            if val == "":
                                continue
                            
                            # Phục hồi thử datetime từ chuỗi ISO nếu có
                            try:
                                # Kiểm tra nhanh nếu chuỗi giống định dạng ISO
                                if len(val) >= 10 and "T" in val or val.count("-") == 2:
                                    parsed_val = datetime.datetime.fromisoformat(val)
                                    ws.cell(row=r_idx, column=c_idx, value=parsed_val)
                                    continue
                            except ValueError:
                                pass
                                
                            ws.cell(row=r_idx, column=c_idx, value=val)

            # Áp dụng Styles từ metadata
            for style_data in sheet_info.get("styles", []):
                r = style_data["r"]
                c = style_data["c"]
                cell = ws.cell(row=r, column=c)

                # Phục hồi Formula
                if "formula" in style_data:
                    cell.value = style_data["formula"]

                # Phục hồi Màu nền
                if "bg" in style_data:
                    bg = style_data["bg"]
                    hex_color = bg if len(bg) == 8 else "FF" + bg
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

                # Phục hồi Font
                if "font" in style_data:
                    font_info = style_data["font"]
                    f_color = font_info.get("color")
                    f_hex = (f_color if len(f_color) == 8 else "FF" + f_color) if f_color else None
                    cell.font = Font(
                        name=font_info.get("name"),
                        size=font_info.get("size"),
                        bold=font_info.get("bold"),
                        italic=font_info.get("italic"),
                        color=f_hex
                    )

                # Phục hồi Căn lề
                if "align" in style_data:
                    align_info = style_data["align"]
                    cell.alignment = Alignment(
                        horizontal=align_info.get("horizontal"),
                        vertical=align_info.get("vertical"),
                        wrap_text=align_info.get("wrap_text")
                    )

                # Phục hồi Border
                if "border" in style_data:
                    border_info = style_data["border"]
                    cell.border = Border(
                        left=Side(style=border_info.get("left")) if border_info.get("left") else None,
                        right=Side(style=border_info.get("right")) if border_info.get("right") else None,
                        top=Side(style=border_info.get("top")) if border_info.get("top") else None,
                        bottom=Side(style=border_info.get("bottom")) if border_info.get("bottom") else None
                    )

                # Phục hồi Định dạng số
                if "num_fmt" in style_data:
                    cell.number_format = style_data["num_fmt"]

                # Phục hồi Hyperlink
                if "link" in style_data:
                    cell.hyperlink = style_data["link"]

                # Phục hồi Comment
                if "comment" in style_data:
                    cell.comment = Comment(style_data["comment"], "Author")

    wb.save(xlsx_path)
    print(f"✅ Đã phục hồi thành công Excel từ gói ZIP tại: {xlsx_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Sử dụng: python zip2xlsx.py <input_data.zip> <output.xlsx>")
        sys.exit(1)
    deserialize_from_zip(sys.argv[1], sys.argv[2])
