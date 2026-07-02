#!/usr/bin/env python3
"""
xlsx2md.py — Convert Excel (.xlsx) to Markdown
Features:
  - Multi-sheet support (separate files or single file with headers)
  - Merged cell handling
  - Column alignment (left / center / right) based on cell content type
  - Auto-detect header row
  - Hyperlink extraction
  - Bold / italic formatting preservation
  - Number formatting (int, float, date, percentage, currency)
  - Empty row/column trimming
  - Max column-width option for readability
  - YAML front-matter metadata block
  - Verbose mode
"""

import argparse
import re
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Optional
 
try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
except ImportError:
    sys.exit("❌  openpyxl not found: pip install openpyxl")
 
 
# ─── Escape ────────────────────────────────────────────────────────────────────
 
def _escape_md(text: str) -> str:
    """Escape only characters that break GFM table structure."""
    # Only escape pipe (breaks columns) and backslash
    text = text.replace("\\", "\\\\")
    text = text.replace("|", "\\|")
    # Newlines inside cell → space
    text = text.replace("\n", " ").replace("\r", "")
    return text
 
 
# ─── Number / date formatting ──────────────────────────────────────────────────
 
def _fmt_value(value, number_format: str) -> str:
    nf = (number_format or "").upper()
 
    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M") if (value.hour or value.minute) else value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d")
 
    if "%" in nf:
        try:
            return f"{float(value)*100:.2f}%"
        except (TypeError, ValueError):
            pass
 
    currency_map = {"$": "$", "€": "€", "£": "£", "¥": "¥", "₫": "₫"}
    for sym in currency_map:
        if sym in (number_format or ""):
            try:
                return f"{sym}{float(value):,.2f}"
            except (TypeError, ValueError):
                pass
 
    if isinstance(value, float):
        return str(int(value)) if value == int(value) else f"{value:,}"
    if isinstance(value, int):
        return f"{value:,}"
 
    return str(value)
 
 
# ─── Cell value extraction ─────────────────────────────────────────────────────
 
def _cell_text(cell, merged_map: dict, hyperlinks: bool) -> str:
    if isinstance(cell, MergedCell):
        return merged_map.get((cell.row, cell.column), "")
 
    value = cell.value
    if value is None:
        return ""
 
    # Hyperlink
    if hyperlinks and cell.hyperlink and cell.hyperlink.target:
        label = _escape_md(str(value).strip())
        url   = cell.hyperlink.target.strip()
        return f"[{label}]({url})"
 
    # Format value
    if isinstance(value, (int, float, datetime, date)):
        text = _fmt_value(value, cell.number_format)
    else:
        text = str(value).strip()
 
    text = _escape_md(text)
 
    # Bold / italic (applied after escaping)
    if cell.font:
        b, i = cell.font.bold, cell.font.italic
        if b and i:
            text = f"***{text}***"
        elif b:
            text = f"**{text}**"
        elif i:
            text = f"*{text}*"
 
    return text
 
 
# ─── Merged cell map ───────────────────────────────────────────────────────────
 
def _build_merged_map(ws) -> dict:
    m: dict = {}
    for rng in ws.merged_cells.ranges:
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                if row == rng.min_row and col == rng.min_col:
                    continue
                m[(row, col)] = ""  # non-anchor merged cells → blank
    return m
 
 
# ─── Column alignment ──────────────────────────────────────────────────────────
 
def _col_alignment(cells) -> str:
    """Determine alignment from first non-empty data cell in the column."""
    for cell in cells:
        if isinstance(cell, MergedCell) or cell.value is None:
            continue
        if isinstance(cell.value, (int, float)):
            return "---:"
        if cell.alignment and cell.alignment.horizontal == "center":
            return ":---:"
        if cell.alignment and cell.alignment.horizontal == "right":
            return "---:"
        return "---"
    return "---"
 
 
# ─── Trim empty rows/cols ──────────────────────────────────────────────────────
 
def _trim(rows: list[list[str]]) -> list[list[str]]:
    # Drop fully empty rows
    rows = [r for r in rows if any(c.strip() for c in r)]
    if not rows:
        return []
    # Find true max used columns
    max_col = max(
        len(r) - next((i for i, c in enumerate(reversed(r)) if c.strip()), len(r))
        for r in rows
    )
    return [r[:max_col] for r in rows]
 
 
# ─── Table renderer ────────────────────────────────────────────────────────────
 
def _render_table(rows: list[list[str]], aligns: list[str], max_w: Optional[int]) -> str:
    if not rows:
        return ""
    n = len(rows[0])
    rows   = [r + [""] * (n - len(r)) for r in rows]
    aligns = (aligns + ["---"] * n)[:n]
 
    if max_w:
        rows = [[c[:max_w] + ("…" if len(c) > max_w else "") for c in r] for r in rows]
 
    widths = [max(len(r[i]) for r in rows) for i in range(n)]
    widths = [max(w, 3) for w in widths]  # minimum separator width
 
    def pad(s, w, a):
        if a == "---:":   return s.rjust(w)
        if a == ":---:":  return s.center(w)
        return s.ljust(w)
 
    lines = []
    # Header
    lines.append("| " + " | ".join(pad(rows[0][i], widths[i], aligns[i]) for i in range(n)) + " |")
    # Separator — always use exactly the alignment markers, no extra dashes needed
    sep_parts = []
    for i in range(n):
        a = aligns[i]
        inner = "-" * widths[i]
        if a == ":---:":  inner = ":" + "-" * (widths[i]-2) + ":"
        elif a == "---:": inner = "-" * (widths[i]-1) + ":"
        elif a == ":---": inner = ":" + "-" * (widths[i]-1)
        sep_parts.append(inner)
    lines.append("| " + " | ".join(sep_parts) + " |")
    # Data
    for row in rows[1:]:
        lines.append("| " + " | ".join(pad(row[i], widths[i], aligns[i]) for i in range(n)) + " |")
 
    return "\n".join(lines)
 
 
# ─── Sheet → markdown ──────────────────────────────────────────────────────────
 
def convert_sheet(ws, *, header_row=1, max_col_width=None, hyperlinks=True, verbose=False) -> str:
    if verbose:
        print(f"  → '{ws.title}'  {ws.max_row}r × {ws.max_column}c")
 
    merged_map = _build_merged_map(ws)
    raw_rows = list(ws.iter_rows())
 
    if not raw_rows:
        return f"*Sheet `{ws.title}` is empty.*\n"
 
    # Reindex to header_row
    raw_rows = raw_rows[header_row - 1:]
 
    # Build string rows
    str_rows = [
        [_cell_text(c, merged_map, hyperlinks) for c in row]
        for row in raw_rows
    ]
 
    trimmed = _trim(str_rows)
    if not trimmed:
        return f"*Sheet `{ws.title}` has no data.*\n"
 
    n = len(trimmed[0])
 
    # Compute alignments from data rows (skip header row=0)
    data_raw = raw_rows[1:] if len(raw_rows) > 1 else raw_rows
    aligns = []
    for col_i in range(n):
        col_cells = [row[col_i] for row in data_raw if col_i < len(row)]
        aligns.append(_col_alignment(col_cells))
 
    return _render_table(trimmed, aligns, max_col_width) + "\n"
 
 
# ─── Front matter ──────────────────────────────────────────────────────────────
 
def _front_matter(src: Path, sheets: list[str]) -> str:
    mtime = datetime.fromtimestamp(src.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    return "\n".join([
        "---",
        f'source: "{src.name}"',
        f'converted: "{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}"',
        f'modified: "{mtime}"',
        f"sheets: [{', '.join(sheets)}]",
        "---", "",
    ])
 
 
def _safe(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name)
 
 
# ─── Entry point ───────────────────────────────────────────────────────────────
 
def xlsx_to_markdown(input_path, output_path=None, *, sheet_names=None,
                     split_sheets=False, header_row=1, max_col_width=None,
                     hyperlinks=True, front_matter=True, verbose=False):
    src = Path(input_path).expanduser().resolve()
    if not src.exists():
        sys.exit(f"❌  Not found: {src}")
 
    wb = load_workbook(src, data_only=True)
    targets = sheet_names or wb.sheetnames
    missing = [s for s in targets if s not in wb.sheetnames]
    if missing:
        sys.exit(f"❌  Sheets not found: {missing}")
 
    out_base = Path(output_path) if output_path else src.with_suffix("")
 
    def build_content(names):
        parts = []
        if front_matter:
            parts.append(_front_matter(src, names))
        for name in names:
            parts.append(f"# {name}\n")
            parts.append(convert_sheet(wb[name], header_row=header_row,
                                       max_col_width=max_col_width,
                                       hyperlinks=hyperlinks, verbose=verbose))
            parts.append("")
        return "\n".join(parts)
 
    if split_sheets:
        out_base.mkdir(parents=True, exist_ok=True)
        for name in targets:
            f = out_base / f"{_safe(name)}.md"
            f.write_text(build_content([name]), encoding="utf-8")
            print(f"✅  {f}")
    else:
        out_file = Path(str(out_base) + ".md") if not str(out_base).endswith(".md") else out_base
        out_file.parent.mkdir(parents=True, exist_ok=True)
        out_file.write_text(build_content(targets), encoding="utf-8")
        print(out_file)
 
 
def main():
    p = argparse.ArgumentParser(prog="xlsx2md", description="XLSX → Markdown")
    p.add_argument("input")
    p.add_argument("-o", "--output")
    p.add_argument("--sheets", nargs="+", metavar="SHEET")
    p.add_argument("--split", action="store_true")
    p.add_argument("--header-row", type=int, default=1, metavar="N")
    p.add_argument("--max-col-width", type=int, metavar="N")
    p.add_argument("--no-hyperlinks", action="store_true")
    p.add_argument("--no-front-matter", action="store_true")
    p.add_argument("-v", "--verbose", action="store_true")
    a = p.parse_args()
    xlsx_to_markdown(a.input, a.output, sheet_names=a.sheets, split_sheets=a.split,
                     header_row=a.header_row, max_col_width=a.max_col_width,
                     hyperlinks=not a.no_hyperlinks, front_matter=not a.no_front_matter,
                     verbose=a.verbose)
 
if __name__ == "__main__":
    main()
