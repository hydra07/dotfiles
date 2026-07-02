import json
import os
import sys
import csv
import io
import zipfile
import datetime
import openpyxl
from openpyxl.utils import get_column_letter

from _xlsx_common import ExcelJSONEncoder, get_hex_color, setup_utf8_stdout

setup_utf8_stdout()


def serialize_to_zip(xlsx_path, zip_output_path):
    print(f"⏳ Đang đọc file Excel: {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)

    metadata = {
        "filename": os.path.basename(xlsx_path),
        "sheets": {}
    }

    # Tạo file ZIP để ghi trực tiếp
    with zipfile.ZipFile(zip_output_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for name in wb.sheetnames:
            print(f"  -> Xử lý sheet: {name}")
            ws = wb[name]
            ws_form = wb_formulas[name]

            # Tìm biên dữ liệu
            active_rows = [1]
            active_cols = [1]
            for (r, c), cell in ws._cells.items():
                has_fill = cell.fill and cell.fill.patternType and cell.fill.patternType != 'none'
                if cell.value is not None or has_fill:
                    active_rows.append(r)
                    active_cols.append(c)

            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col_idx, max_row_idx = merged_range.bounds
                active_rows.extend([min_row, max_row_idx])
                active_cols.extend([min_col, max_col_idx])

            max_row = max(active_rows)
            max_col = max(active_cols)

            sheet_info = {
                "show_grid_lines": bool(ws.views.sheetView[0].showGridLines) if ws.views.sheetView else True,
                "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
                "col_widths": {},
                "row_heights": {},
                "styles": []
            }

            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                width = ws.column_dimensions[col_letter].width
                if width is not None:
                    sheet_info["col_widths"][col_idx] = width

            for row_idx in range(1, max_row + 1):
                height = ws.row_dimensions[row_idx].height
                if height is not None:
                    sheet_info["row_heights"][row_idx] = height

            # Thu thập dữ liệu CSV và Style
            csv_buffer = io.StringIO()
            csv_writer = csv.writer(csv_buffer)

            for r in range(1, max_row + 1):
                csv_row = []
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    
                    # 1. Ghi giá trị vào dòng CSV
                    val = cell.value
                    if isinstance(val, (datetime.datetime, datetime.date, datetime.time)):
                        val = val.isoformat()
                    csv_row.append(val if val is not None else "")

                    # 2. Thu thập Style (Màu, font, border, formula) nếu có
                    has_style = False
                    style_data = {"r": r, "c": c}

                    cell_form = ws_form.cell(row=r, column=c)
                    formula = str(cell_form.value) if cell_form.value and str(cell_form.value).startswith("=") else None
                    if formula:
                        style_data["formula"] = formula
                        has_style = True

                    has_fill = cell.fill and cell.fill.patternType and cell.fill.patternType != 'none'
                    if has_fill:
                        bg_color = get_hex_color(cell.fill.fgColor)
                        if bg_color:
                            style_data["bg"] = bg_color
                            has_style = True

                    if cell.font and (cell.font.bold or cell.font.italic or cell.font.color or cell.font.name != "Calibri"):
                        style_data["font"] = {
                            "name": cell.font.name,
                            "size": cell.font.size,
                            "bold": cell.font.bold,
                            "italic": cell.font.italic,
                            "color": get_hex_color(cell.font.color)
                        }
                        has_style = True

                    if cell.alignment and (cell.alignment.horizontal or cell.alignment.vertical or cell.alignment.wrap_text):
                        style_data["align"] = {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": cell.alignment.wrap_text
                        }
                        has_style = True

                    if cell.border and (cell.border.left.style or cell.border.right.style or cell.border.top.style or cell.border.bottom.style):
                        style_data["border"] = {
                            "left": cell.border.left.style,
                            "right": cell.border.right.style,
                            "top": cell.border.top.style,
                            "bottom": cell.border.bottom.style
                        }
                        has_style = True

                    if cell.number_format and cell.number_format != 'General':
                        style_data["num_fmt"] = cell.number_format
                        has_style = True
                    if hasattr(cell, 'hyperlink') and cell.hyperlink and cell.hyperlink.target:
                        style_data["link"] = cell.hyperlink.target
                        has_style = True
                    if cell.comment:
                        style_data["comment"] = cell.comment.text
                        has_style = True

                    if has_style:
                        sheet_info["styles"].append(style_data)

                # Dọn dòng CSV ở cuối nếu toàn rỗng để tránh thừa thãi
                while csv_row and csv_row[-1] == "":
                    csv_row.pop()
                csv_writer.writerow(csv_row)

            # Lưu cấu hình style của sheet
            metadata["sheets"][name] = sheet_info

            # Ghi sheet CSV vào trong ZIP
            zf.writestr(f"sheets/{name}.csv", csv_buffer.getvalue().encode('utf-8-sig'))

        # Ghi metadata JSON vào ZIP
        json_bytes = json.dumps(metadata, ensure_ascii=False, indent=2, cls=ExcelJSONEncoder).encode('utf-8')
        zf.writestr("metadata.json", json_bytes)

    print(f"✅ Đã đóng gói toàn bộ Data (CSV) và Metadata (JSON) vào: {zip_output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Sử dụng: python xlsx2zip.py <input.xlsx> <output_data.zip>")
        sys.exit(1)
    serialize_to_zip(sys.argv[1], sys.argv[2])
