import json
import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter

from _xlsx_common import ExcelJSONEncoder, get_hex_color, setup_utf8_stdout

setup_utf8_stdout()


def serialize_workbook(xlsx_path):
    print(f"⏳ Đang đọc toàn bộ file Excel: {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)

    wb_data = {
        "filename": os.path.basename(xlsx_path),
        "sheets": {}
    }

    for name in wb.sheetnames:
        print(f"  -> Đang quét sheet: {name}")
        ws = wb[name]
        ws_form = wb_formulas[name]

        # Quét giới hạn thực tế để tối ưu kích thước file JSON
        active_rows = [1]
        active_cols = [1]
        for (r, c), cell in ws._cells.items():
            # Chỉ lấy các ô có giá trị hoặc được tô màu nền/chữ hoặc định dạng đặc biệt
            has_fill = cell.fill and cell.fill.patternType and cell.fill.patternType != 'none'
            if cell.value is not None or has_fill:
                active_rows.append(r)
                active_cols.append(c)

        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col_idx, max_row_idx = merged_range.bounds
            active_rows.extend([min_row, max_row_idx])
            active_cols.extend([min_col, max_col_idx])

        max_row = max(active_rows)
        max_col = max(active_cols)

        sheet_info = {
            "show_grid_lines": bool(ws.views.sheetView[0].showGridLines) if ws.views.sheetView else True,
            "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
            "col_widths": {},
            "row_heights": {},
            "cells": []
        }

        # Lưu độ rộng cột
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            if width is not None:
                sheet_info["col_widths"][col_idx] = width

        # Lưu độ cao dòng
        for row_idx in range(1, max_row + 1):
            height = ws.row_dimensions[row_idx].height
            if height is not None:
                sheet_info["row_heights"][row_idx] = height

        # Quét các ô dữ liệu (Sử dụng cấu trúc Sparse để tiết kiệm dung lượng)
        for (r, c), cell in ws._cells.items():
            # Bỏ qua các ô không có giá trị VÀ không có màu nền (các ô rỗng chỉ có định dạng font sẽ bị bỏ qua để giảm thiểu dung lượng)
            has_fill = cell.fill and cell.fill.patternType and cell.fill.patternType != 'none'
            if cell.value is None and not has_fill:
                continue

            cell_form = ws_form.cell(row=r, column=c)
            
            # Formula
            formula = str(cell_form.value) if cell_form.value and str(cell_form.value).startswith("=") else None

            # Style màu nền
            bg_color = None
            if has_fill:
                bg_color = get_hex_color(cell.fill.fgColor)

            # Font style
            font_info = None
            if cell.font and (cell.font.bold or cell.font.italic or cell.font.color or cell.font.name != "Calibri"):
                font_info = {
                    "name": cell.font.name,
                    "size": cell.font.size,
                    "bold": cell.font.bold,
                    "italic": cell.font.italic,
                    "color": get_hex_color(cell.font.color)
                }

            # Alignment
            align_info = None
            if cell.alignment and (cell.alignment.horizontal or cell.alignment.vertical or cell.alignment.wrap_text):
                align_info = {
                    "horizontal": cell.alignment.horizontal,
                    "vertical": cell.alignment.vertical,
                    "wrap_text": cell.alignment.wrap_text
                }

            # Border
            border_info = None
            if cell.border and (cell.border.left.style or cell.border.right.style or cell.border.top.style or cell.border.bottom.style):
                border_info = {
                    "left": cell.border.left.style,
                    "right": cell.border.right.style,
                    "top": cell.border.top.style,
                    "bottom": cell.border.bottom.style
                }

            # Loại bỏ các key không có dữ liệu để tiết kiệm tối đa dung lượng JSON
            cell_data = {"r": r, "c": c}
            if cell.value is not None:
                cell_data["val"] = cell.value
            if formula:
                cell_data["formula"] = formula
            if bg_color:
                cell_data["bg"] = bg_color
            if font_info:
                cell_data["font"] = font_info
            if align_info:
                cell_data["align"] = align_info
            if border_info:
                cell_data["border"] = border_info
            
            # Format số / Hyperlink / Comment
            if cell.number_format and cell.number_format != 'General':
                cell_data["num_fmt"] = cell.number_format
            if hasattr(cell, 'hyperlink') and cell.hyperlink and cell.hyperlink.target:
                cell_data["link"] = cell.hyperlink.target
            if cell.comment:
                cell_data["comment"] = cell.comment.text

            sheet_info["cells"].append(cell_data)

        wb_data["sheets"][name] = sheet_info

    return wb_data

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Sử dụng: python xlsx2json_full.py <input.xlsx> <output.json>")
        sys.exit(1)

    serialize_data = serialize_workbook(sys.argv[1])
    with open(sys.argv[2], "w", encoding="utf-8") as f:
        json.dump(serialize_data, f, ensure_ascii=False, indent=2, cls=ExcelJSONEncoder)
    print(f"✅ Đã lưu toàn bộ dữ liệu workbook vào: {sys.argv[2]}")
