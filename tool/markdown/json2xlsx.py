import json
import os
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ngược lại với COLOR_MAP trong xlsx2md.py
COLOR_HEX_MAP = {
    "gray": "FFC0C0C0",
    "purple": "FF7030A0",
    "blue": "FF5B9BD5",
    "red": "FFFF6B6B",
    "green": "FF70AD47",
    "orange": "FFFF6E05",
    "yellow": "FFFFF5A4",
    "pink": "FFFFDDEB",
    "tan": "FFFFB572",
    "light-red": "FFFFBFBF",
    "cream": "FFFFE6CA",
}

# Các tuần ứng với từng tháng cho File 2 (FY26)
MONTH_TO_WEEKS = {
    "Apr-26": [14, 15, 16, 17],
    "May-26": [18, 19, 20, 21],
    "Jun-26": [22, 23, 24, 25, 26],
    "Jul-26": [27, 28, 29, 30],
    "Aug-26": [31, 32, 33, 34, 35],
    "Sep-26": [36, 37, 38, 39],
    "Oct-26": [40, 41, 42, 43, 44],
    "Nov-26": [45, 46, 47, 48],
    "Dec-26": [49, 50, 51, 52, 53],
    "Jan-27": [1, 2, 3, 4],
    "Feb-27": [5, 6, 7, 8],
    "Mar-27": [9, 10, 11, 12, 13],
}

def create_projects_xlsx(projects_data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FY26スケジュール・リソース管理表"
    
    # Kích hoạt gridlines
    ws.views.sheetView[0].showGridLines = True

    # Font mặc định
    font_main = Font(name="Segoe UI", size=10)
    font_header = Font(name="Segoe UI", size=10, bold=True)
    font_title = Font(name="Segoe UI", size=14, bold=True)

    # Style viền
    thin = Side(border_style="thin", color="D3D3D3")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Tiêu đề bảng
    ws.cell(row=2, column=1, value="FY26年間スケジュール・リソースアサイン一覧").font = font_title

    # Header của bảng thông tin
    headers_info = [
        "案件情報", "プロジェクトID", "お客様名", "案件種別", 
        "希望時期", "車両", "DP", "運行管理", "備考", "更新日"
    ]
    for col_idx, h in enumerate(headers_info, 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all
        ws.row_dimensions[5].height = 28

    # Tạo header lịch tuần cho biểu đồ Gantt
    # Định dạng các cột tuần từ cột 13 (M) trở đi
    current_col = 13
    week_col_map = {} # map (week_num) -> col_idx

    # Ghi Month Header ở dòng 4 và Week Header ở dòng 5
    for month, weeks in MONTH_TO_WEEKS.items():
        start_col = current_col
        for w in weeks:
            # Ghi số tuần ở dòng 5
            cell_w = ws.cell(row=5, column=current_col, value=w)
            cell_w.font = font_header
            cell_w.alignment = Alignment(horizontal="center", vertical="center")
            cell_w.border = border_all
            week_col_map[(month, w)] = current_col
            current_col += 1
        
        # Merge các cột của tháng đó lại ở dòng 4
        end_col = current_col - 1
        ws.merge_cells(start_row=4, end_row=4, start_column=start_col, end_column=end_col)
        cell_m = ws.cell(row=4, column=start_col, value=month)
        cell_m.font = font_header
        cell_m.alignment = Alignment(horizontal="center", vertical="center")
        cell_m.border = border_all

    # Định dạng độ rộng cột sơ bộ
    for c in range(1, 11):
        ws.column_dimensions[get_column_letter(c)].width = 15
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['I'].width = 35

    for c in range(13, current_col):
        ws.column_dimensions[get_column_letter(c)].width = 4

    # Ghi dữ liệu các projects
    current_row = 6
    current_section = None

    for p in projects_data:
        # Nếu đổi Section, tạo dòng phân cách Section giống Excel gốc
        if p.get("section") != current_section:
            current_section = p["section"]
            ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=current_col-1)
            sec_cell = ws.cell(row=current_row, column=1, value=current_section)
            sec_cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            sec_cell.fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")
            sec_cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[current_row].height = 24
            current_row += 1

        # Điền thông tin cơ bản của project
        fields = [
            ("project_id", 2),
            ("customer", 3),
            ("project_type", 4),
            ("period", 5),
            ("vehicles", 6),
            ("dp", 7),
            ("monitoring", 8),
            ("notes", 9),
            ("last_updated", 10),
        ]
        
        for key, col in fields:
            val = p.get(key, "")
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.font = font_main
            cell.border = border_all
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Đặt lại cột 1 (Section) ở mỗi dòng cho rõ ràng
        sec_lbl = ws.cell(row=current_row, column=1, value=current_section)
        sec_lbl.font = font_main
        sec_lbl.border = border_all

        # Điền dữ liệu Gantt (màu sắc + text/milestone)
        schedule = p.get("schedule", {})
        for month, entry in schedule.items():
            if month not in MONTH_TO_WEEKS:
                continue
            
            weeks = MONTH_TO_WEEKS[month]
            colors = entry.get("colors", [])
            labels = entry.get("labels", [])
            milestones = entry.get("milestones", [])

            # Ghép nhãn và mốc thời gian
            text_items = milestones + labels
            
            # Phân bổ nội dung/màu sắc vào các tuần của tháng đó
            for idx, w in enumerate(weeks):
                col_idx = week_col_map.get((month, w))
                if not col_idx:
                    continue
                
                cell_gantt = ws.cell(row=current_row, column=col_idx)
                cell_gantt.border = border_all

                # Tô màu (phân bổ màu đều cho các tuần)
                if colors:
                    color_name = colors[idx % len(colors)]
                    hex_color = COLOR_HEX_MAP.get(color_name)
                    # Nếu là mã màu tùy chỉnh
                    if not hex_color and len(color_name) == 6:
                        hex_color = "FF" + color_name
                    
                    if hex_color:
                        cell_gantt.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

                # Ghi text (phân bổ nhãn vào các tuần tương ứng)
                if idx < len(text_items):
                    cell_gantt.value = text_items[idx]
                    cell_gantt.font = Font(name="Segoe UI", size=8, bold=True)
                    cell_gantt.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[current_row].height = 20
        current_row += 1

    wb.save(output_path)
    print(f"✅ Đã tạo thành công file Excel phục hồi tại: {output_path}")

def create_areas_xlsx(areas_data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet2"
    ws.views.sheetView[0].showGridLines = True

    font_main = Font(name="Segoe UI", size=10)
    font_header = Font(name="Segoe UI", size=10, bold=True)
    thin = Side(border_style="thin", color="D3D3D3")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Viết tiêu đề cột
    ws.cell(row=3, column=2, value="EVO実証実験エリア").font = font_header
    ws.cell(row=3, column=2).border = border_all

    # Thu thập tất cả các tháng xuất hiện trong dữ liệu để tạo cột
    all_months = set()
    for area in areas_data:
        all_months.update(area.get("schedule", {}).keys())
    
    # Sắp xếp các tháng theo trình tự thời gian (ví dụ: 2026/4, 2026/5...)
    def sort_key(m_str):
        try:
            parts = m_str.split("/")
            return int(parts[0]), int(parts[1])
        except Exception:
            return 9999, 99

    sorted_months = sorted(list(all_months), key=sort_key)

    # Viết header tháng
    month_col_map = {}
    for col_idx, m_str in enumerate(sorted_months, 3):
        # Tách năm và tháng
        parts = m_str.split("/")
        year = parts[0] if len(parts) > 0 else ""
        month = parts[1] if len(parts) > 1 else ""

        # Ghi năm ở dòng 3
        cell_y = ws.cell(row=3, column=col_idx, value=year)
        cell_y.font = font_header
        cell_y.alignment = Alignment(horizontal="center")
        cell_y.border = border_all

        # Ghi tháng ở dòng 4
        cell_m = ws.cell(row=4, column=col_idx, value=month)
        cell_m.font = font_header
        cell_m.alignment = Alignment(horizontal="center")
        cell_m.border = border_all
        
        month_col_map[m_str] = col_idx

    ws.column_dimensions['B'].width = 25
    for c in range(3, len(sorted_months) + 3):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Viết dữ liệu khu vực
    current_row = 5
    for area in areas_data:
        area_name = area.get("area_name", "")
        cell_name = ws.cell(row=current_row, column=2, value=area_name)
        cell_name.font = font_main
        cell_name.border = border_all

        schedule = area.get("schedule", {})
        for m_str, entry in schedule.items():
            col_idx = month_col_map.get(m_str)
            if not col_idx:
                continue

            cell_sch = ws.cell(row=current_row, column=col_idx)
            cell_sch.border = border_all

            # Điền text nếu có
            if "text" in entry:
                cell_sch.value = entry["text"]
                cell_sch.font = font_main
                cell_sch.alignment = Alignment(horizontal="center", vertical="center")

            # Điền màu nếu có
            color_name = entry.get("color")
            hex_color = COLOR_HEX_MAP.get(color_name) if color_name else None
            if not hex_color and "color_raw" in entry:
                hex_color = "FF" + entry["color_raw"]

            if hex_color:
                cell_sch.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        
        current_row += 1

    wb.save(output_path)
    print(f"✅ Đã tạo thành công file Excel phục hồi tại: {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Sử dụng: python json2xlsx.py <file_json_dau_vao> <file_excel_dau_ra>")
        print("Ví dụ:")
        print("  python json2xlsx.py file2_projects.json file2_restored.xlsx")
        sys.exit(1)

    json_path = sys.argv[1]
    xlsx_path = sys.argv[2]

    if not os.path.exists(json_path):
        print(f"❌ Lỗi: Không tìm thấy file JSON đầu vào '{json_path}'")
        sys.exit(1)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Tự động phát hiện cấu trúc dữ liệu để gọi hàm phù hợp
    if isinstance(data, list) and len(data) > 0:
        first_item = data[0]
        if "project_id" in first_item:
            print("Detected: FY26 Projects list. Rebuilding...")
            create_projects_xlsx(data, xlsx_path)
        elif "area_name" in first_item:
            print("Detected: EVO Areas list. Rebuilding...")
            create_areas_xlsx(data, xlsx_path)
        else:
            print("❌ Lỗi: Cấu trúc JSON không được hỗ trợ để chuyển đổi ngược lại.")
    else:
        print("❌ Lỗi: Dữ liệu JSON phải là danh sách các phần tử (Array).")
